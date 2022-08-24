import numpy as np
import openpyxl as op


inp = op.load_workbook("input.xlsx", data_only=True).active
if inp.cell(row=2, column=1).value == 'Radial':
    Task = 1
else:
    Task = 2
R = inp.cell(row=2, column=2).value
H = inp.cell(row=2, column=3).value
if Task == 1:
    h = inp.cell(row=2, column=4).value
a = inp.cell(row=1, column=12).value
dt = np.zeros(a-2)
d = np.zeros(a-2)
for i in range(2, a):
    dt[i-2] = int(inp.cell(row=i, column=12).value)
    d[i-2] = inp.cell(row=i, column=6).value
a = inp.cell(row=2, column=7).value
E = np.zeros(inp.max_row-1)
q = np.zeros(inp.max_row-1)
A = np.zeros(inp.max_row-1)
for i in range(2, inp.max_row+1):
    E[i-2] = inp.cell(row=i, column=9).value
    q[i-2] = inp.cell(row=i, column=10).value
    A[i-2] = inp.cell(row=i, column=11).value