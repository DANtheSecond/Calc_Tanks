import numpy as np
from scipy.interpolate import CubicSpline
import openpyxl as op
from mumatrix import tab_p, tab_b, tab_k, tab_musR, tab_E, tab_mu, tab_G, tab_musH, tab_b1, tab_aH, tab_RH, tab_Z
from CoeffConv import tab_En, tab_Ga
from datetime import datetime

# Functions
def ApproxPov(x, y, x1):
    if y[0] == 0:
        return 0
    if x[0] == 0:
        ones = np.ones(len(x))
        A0 = np.array([[0., 0.],
                       [0., 0.]])
        A1 = np.array([[0., 0.],
                       [0., 0.]])
        A0[0, 0] = np.sum(np.log(y))
        A0[0, 1] = np.sum(np.log(x + ones))
        A0[1, 0] = np.sum(np.log(y) * np.log(x + ones))
        A1[0, 0] = float(len(x))
        A1[0, 1] = np.sum(np.log(x + ones))
        A1[1, 0] = np.sum(np.log(x + ones))
        for i in range(len(x)):
            A0[1, 1] += np.log(x[i] + 1) ** 2
            A1[1, 1] += np.log(x[i] + 1) ** 2
        a = np.exp(np.linalg.det(A0) / np.linalg.det(A1))
        A0[0, 0] = float(len(x))
        A0[0, 1] = np.sum(np.log(y))
        A0[1, 0] = np.sum(np.log(x + ones))
        A0[1, 1] = np.sum(np.log(y) * np.log(x + ones))
        A1[0, 0] = float(len(x))
        A1[0, 1] = np.sum(np.log(x + ones))
        A1[1, 0] = np.sum(np.log(x + ones))
        A1[1, 1] = 0
        for i in range(len(x)):
            A1[1, 1] += np.log(x[i] + 1) ** 2
        b = np.linalg.det(A0) / np.linalg.det(A1)
        c = (y[0] / a) ** (-b)
        return a * (x1 + c) ** b
    else:
        A0 = np.array([[0., 0.],
                       [0., 0.]])
        A1 = np.array([[0., 0.],
                       [0., 0.]])
        A0[0, 0] = np.sum(np.log(y))
        A0[0, 1] = np.sum(np.log(x))
        A0[1, 0] = np.sum(np.log(y) * np.log(x))
        A1[0, 0] = float(len(x))
        A1[0, 1] = np.sum(np.log(x))
        A1[1, 0] = np.sum(np.log(x))
        for i in range(len(x)):
            A0[1, 1] += np.log(x[i]) ** 2
            A1[1, 1] += np.log(x[i]) ** 2
        a = np.exp(np.linalg.det(A0) / np.linalg.det(A1))
        A0[0, 0] = float(len(x))
        A0[0, 1] = np.sum(np.log(y))
        A0[1, 0] = np.sum(np.log(x))
        A0[1, 1] = np.sum(np.log(y) * np.log(x))
        A1[0, 0] = float(len(x))
        A1[0, 1] = np.sum(np.log(x))
        A1[1, 0] = np.sum(np.log(x))
        A1[1, 1] = 0
        for i in range(len(x)):
            A1[1, 1] += np.log(x[i]) ** 2
        b = np.linalg.det(A0) / np.linalg.det(A1)
        return a * x1 ** b

def ApproxLog(x, y, x1):
    if x[0] == 0:
        ones = np.ones(len(x))
        A0 = np.array([[0., 0.],
                       [0., 0.]])
        A1 = np.array([[0., 0.],
                       [0., 0.]])
        A0[0, 0] = np.sum(y)
        A0[0, 1] = np.sum(np.log(x+ones))
        A0[1, 0] = np.sum(y * np.log(x+ones))
        A1[0, 0] = float(len(x))
        A1[0, 1] = np.sum(np.log(x+ones))
        A1[1, 0] = np.sum(np.log(x+ones))
        for i in range(len(x)):
            A0[1, 1] += np.log(x[i]+1) ** 2
            A1[1, 1] += np.log(x[i]+1) ** 2
        a = np.linalg.det(A0) / np.linalg.det(A1)
        A0[0, 0] = float(len(x))
        A0[0, 1] = np.sum(y)
        A0[1, 0] = np.sum(np.log(x+ones))
        A0[1, 1] = np.sum(y * np.log(x+ones))
        A1[0, 0] = float(len(x))
        A1[0, 1] = np.sum(np.log(x+ones))
        A1[1, 0] = np.sum(np.log(x+ones))
        A1[1, 1] = 0
        for i in range(len(x)):
            A1[1, 1] += np.log(x[i]+1) ** 2
        b = np.linalg.det(A0) / np.linalg.det(A1)
        c = np.exp((y[0] - a) / b)
        print(a, b, c)
        return a + b * np.log(x1 + c)
    else:
        A0 = np.array([[0., 0.],
                       [0., 0.]])
        A1 = np.array([[0., 0.],
                       [0., 0.]])
        A0[0, 0] = np.sum(y)
        A0[0, 1] = np.sum(np.log(x))
        A0[1, 0] = np.sum(y * np.log(x))
        A1[0, 0] = float(len(x))
        A1[0, 1] = np.sum(np.log(x))
        A1[1, 0] = np.sum(np.log(x))
        for i in range(len(x)):
            A0[1, 1] += np.log(x[i]) ** 2
            A1[1, 1] += np.log(x[i]) ** 2
        a = np.linalg.det(A0) / np.linalg.det(A1)
        A0[0, 0] = float(len(x))
        A0[0, 1] = np.sum(y)
        A0[1, 0] = np.sum(np.log(x))
        A0[1, 1] = np.sum(y * np.log(x))
        A1[0, 0] = float(len(x))
        A1[0, 1] = np.sum(np.log(x))
        A1[1, 0] = np.sum(np.log(x))
        A1[1, 1] = 0
        for i in range(len(x)):
            A1[1, 1] += np.log(x[i]) ** 2
        b = np.linalg.det(A0) / np.linalg.det(A1)
        return a + b * np.log(x1)

def ApproxLin(x, y, x1):
    if x1 > x[-1]:
        x0 = x[-1]
        x2 = x[-2]
    else:
        x0 = x[0]
        x2 = x[1]
        for i in range(len(x) - 1):
            if x1 > x[i]:
                x0 = x[i]
                x2 = x[i + 1]
            else:
                break
    return y[np.where(x == x0)] + (y[np.where(x == x2)] - y[np.where(x == x0)]) * (x1 - x0) / (x2 - x0)

def ApproxExp(x, y, x1):
    if y[0] == 0:
        return 0
    A0 = np.array([[0., 0.],
                   [0., 0.]])
    A1 = np.array([[0., 0.],
                   [0., 0.]])
    A0[0, 0] = np.sum(np.log(y))
    A0[0, 1] = np.sum(x)
    A0[1, 0] = np.sum(np.log(y) * x)
    A1[0, 0] = float(len(x))
    A1[0, 1] = np.sum(x)
    A1[1, 0] = np.sum(x)
    for i in range(len(x)):
        A0[1, 1] += x[i] ** 2
        A1[1, 1] += x[i] ** 2
    a = np.exp(np.linalg.det(A0) / np.linalg.det(A1))
    A0[0, 0] = float(len(x))
    A0[0, 1] = np.sum(np.log(y))
    A0[1, 0] = np.sum(x)
    A0[1, 1] = np.sum(np.log(y) * x)
    A1[0, 0] = float(len(x))
    A1[0, 1] = np.sum(x)
    A1[1, 0] = np.sum(x)
    A1[1, 1] = 0
    for i in range(len(x)):
        A1[1, 1] += x[i] ** 2
    b = np.linalg.det(A0) / np.linalg.det(A1)
    return (a * np.exp(b * x1))

approx_funcs = {
    'exp': ApproxExp,
    'pov': ApproxPov,
    'log': ApproxLog,
    'lin': ApproxLin
}
def Approx(x, y, x1, ap):
    ap = approx_funcs.get(ap)
    if x1 < x[0]:
        if ApproxLin(x, y, x[0]) >= ap(x, y, x[0]):
            return max(ApproxLin(x, y, x1), ap(x, y, x1))
        else:
            if y[0] < y[-1]:  # ascending
                weight = (x1 - x[0]) / (x[0] - x[1])
                if x1 < x[0] + x[0] - x[1]:
                    return ApproxPov(x, y, x1)
                else:
                    return weight * ap(x, y, x1) + (1 - weight) * ApproxLin(x, y, x1)
            else:  # descending
                return min(y[-1], ap(x, y, x1))
    elif x[0] <= x1 <= x[-1]:
        return ApproxLin(x, y, x1)
    else:
        if ApproxLin(x, y, x[-1]) >= ap(x, y, x[-1]):
            return max(ApproxLin(x, y, x1), ap(x, y, x1))
        else:
            if y[0] < y[-1]:  # ascending
                weight = (x1 - x[-1]) / (x[-1] - x[-2])
                if x1 > x[-1] + x[-1] - x[-2]:
                    return ap(x, y, x1)
                else:
                    return weight * ap(x, y, x1) + (1 - weight) * ApproxLin(x, y, x1)
            else:  # descending
                return min(y[-1], ap(x, y, x1))

# Input Data
inp = op.load_workbook("input.xlsx", data_only=True)
N = len(inp.worksheets)-1
results = op.Workbook()
res = results.active
res['S2'] = 'Total dose, μSv/h'
res['T2'] = "=SUM(Q:Q)"
T = int(2)
for t in range(N):
    sht = 'Tank'+str(t+1)
    print(sht)
    inp = op.load_workbook("input.xlsx", data_only=True)[sht]
    rows = len([cell for cell in inp['I'] if cell.value])
    if inp.cell(row=2, column=1).value == 'Radial':
        Task = 1
    else:
        Task = 2
    R = inp.cell(row=2, column=2).value
    H = inp.cell(row=2, column=3).value
    if Task == 1:
        h = inp.cell(row=2, column=4).value
    a = len([cell for cell in inp['F'] if cell.value]) - 1
    dt = np.zeros(a).astype(int)
    d = np.zeros(a)
    for i in range(2, a + 2):
        dt[i - 2] = inp.cell(row=i, column=5).value
        d[i - 2] = inp.cell(row=i, column=6).value
    a = inp.cell(row=2, column=7).value
    E = np.zeros(rows - 1)
    q = np.zeros(rows - 1)
    A = np.zeros(rows - 1)
    for i in range(2, rows + 1):
        E[i - 2] = inp.cell(row=i, column=9).value
        q[i - 2] = inp.cell(row=i, column=10).value
        A[i - 2] = inp.cell(row=i, column=11).value


# Creating workbook
    res.cell(row=T-1, column=1).value = 'Tank №' + str(t+1)
    res.cell(row=T, column=1).value = 'Energy, E, MeV'
    res.cell(row=T, column=2).value = 'Gamma emission, q, rel.un.'
    res.cell(row=T, column=3).value = 'Г, μSv*m2/(h*Bq)'
    res.cell(row=T, column=4).value = 'μs, cm-1'
    res.cell(row=T, column=5).value = 'Protection material (1 - Concrete; 2 - Lead; 3 - Iron)'
    res.cell(row=T, column=6).value = 'Protection thickness, δ, mm'
    for i in range(2, len(d) + 2):
        res.cell(row=T+i-1, column=5).value = dt[i - 2]
        res.cell(row=T+i-1, column=6).value = d[i - 2]
    res.cell(row=T, column=7).value = 'Protection, μ, cm-1'
    if Task == 1:
        res.cell(row=T, column=8).value = 'p, rel.un.'
        res.cell(row=T, column=9).value = "b, rel.un."
        res.cell(row=T, column=10).value = "k', rel.un."
        res.cell(row=T, column=11).value = 'k", rel.un.'
        res.cell(row=T, column=12).value = 'μsR, rel.un.'
        res.cell(row=T, column=13).value = "G', rel.un."
        res.cell(row=T, column=14).value = 'G", rel.un.'
    else:
        res.cell(row=T, column=8).value = 'μsh, rel.un.'
        res.cell(row=T, column=9).value = "b, rel.un."
        res.cell(row=T, column=10).value = "a/h, rel.un."
        res.cell(row=T, column=12).value = 'R/h, rel.un.'
        res.cell(row=T, column=13).value = "Z, rel.un."
    res.cell(row=T, column=15).value = 'D, μSv/h'
    res.cell(row=T, column=16).value = 'Dose from Tank №' + str(t+1) + ', μSv/h'

    # Calculations
    D = np.zeros(len(E))
    if Task == 1:
        p = a / R
        k1 = (H - h) / R
        k2 = h / R
        res.cell(row=T + 1, column=8).value = p
        res.cell(row=T + 1, column=10).value = k1
        res.cell(row=T + 1, column=11).value = k2
        for e in range(len(E)):
            res.cell(row=T + e + 1, column=1).value = E[e]
            res.cell(row=T + e + 1, column=2).value = q[e]
            res.cell(row=T + e + 1, column=3).value = float(ApproxLin(tab_En, tab_Ga, E[e]) * q[e])
            res.cell(row=T + e + 1, column=7).value = ''
            b = 0.
            mud = np.zeros(len(dt))
            if E[e] > tab_E.max() or E[e] < tab_E.min():
                f = CubicSpline(tab_E, tab_mu[:, 0], extrapolate=True)
                if f(E[e]) > 0:
                    mus = f(E[e])
                else:
                    mus = ApproxPov(tab_E, tab_mu[:, 0], E[e])
            else:
                mus = ApproxLin(tab_E, tab_mu[:, 0], E[e])
            res.cell(row=T + e + 1, column=4).value = float(mus)
            print('μs =', mus)
            for i in range(len(dt)):
                if E[e] > tab_E.max() or E[e] < tab_E.min():
                    f = CubicSpline(tab_E, tab_mu[:, dt[i]], extrapolate=True)
                    if f(E[e]) > 0:
                        mud[i] = f(E[e])
                    else:
                        mud[i] = ApproxPov(tab_E, tab_mu[:, dt[i]], E[e])
                else:
                    mud[i] = ApproxLin(tab_E, tab_mu[:, dt[i]], E[e])
                res.cell(row=T + e + 1, column=7).value += ' μd' + str(i+1) + ' = ' + str(mud[i])
                print('μd[', i, '] =', mud[i])
                b += (mud[i] * d[i]) / 10
            res.cell(row=T + e + 1, column=9).value = b
            musR = float(mus * R * 100)
            res.cell(row=T + e + 1, column=12).value = musR
            s = np.zeros(len(tab_musR))
            Gpbk = np.zeros((len(tab_k), len(tab_b), len(tab_p)))
            for i in range(len(tab_p)):
                for j in range(len(tab_b)):
                    for n in range(len(tab_k)):
                        for m in range(len(tab_musR)):
                            s[m] = tab_G[m, n, j, i]
                        Gpbk[n, j, i] = Approx(tab_musR, s, musR, 'pov')
            s = np.zeros(len(tab_k))
            Gpb1 = np.zeros((len(tab_b), len(tab_p)))
            Gpb2 = np.zeros((len(tab_b), len(tab_p)))
            for i in range(len(tab_p)):
                for j in range(len(tab_b)):
                    for n in range(len(tab_k)):
                        s[n] = Gpbk[n, j, i]
                    if k1 > tab_k[-1]:
                        Gpb1[j, i] = tab_k[-1]
                    else:
                        Gpb1[j, i] = ApproxLin(tab_k, s, k1)
                    if k2 > tab_k[-1]:
                        Gpb2[j, i] = tab_k[-1]
                    else:
                        Gpb2[j, i] = ApproxLin(tab_k, s, k2)
            s = np.zeros(len(tab_b))
            Gp1 = np.zeros(len(tab_p))
            Gp2 = np.zeros(len(tab_p))
            for i in range(len(tab_p)):
                for j in range(len(tab_b)):
                    s[j] = Gpb1[j, i]
                Gp1[i] = Approx(tab_b, s, b, 'exp')
                for j in range(len(tab_b)):
                    s[j] = Gpb2[j, i]
                Gp2[i] = Approx(tab_b, s, b, 'exp')
            if k1 == 0:
                G1 = 0
            else:
                G1 = float(Approx(tab_p, Gp1, p, 'pov'))
            res.cell(row=T + e + 1, column=13).value = G1
            if k2 == 0:
                G2 = 0
            else:
                G2 = float(Approx(tab_p, Gp2, p, 'pov'))
            res.cell(row=T + e + 1, column=14).value = G2
            V = H * np.pi * R ** 2
            D[e] = 2 * A[e] * ApproxLin(tab_En, tab_Ga, E[e]) * q[e] * R / V * (G1 + G2)
            res.cell(row=T + e + 1, column=15).value = D[e]
            print('p = ', p)
            print('b = ', b)
            print('μsR = ', musR)
            print("k' = ", k1, "| k'' = ", k2)
            print("G' = ", G1, "| G'' = ", G2)
            print('E = ', E[e])
            print('q = ', q[e])
            print('Г = ', float(ApproxLin(tab_En, tab_Ga, E[e]) * q[e]))
            print('V = ', V)
            print('D[', e, ']= ', D[e], 'μSv/h')
    else:
        aH = a/H
        RH = R/H
        res.cell(row=T + 1, column=10).value = aH
        res.cell(row=T + 1, column=12).value = RH
        for e in range(len(E)):
            res.cell(row=T + e + 1, column=1).value = E[e]
            res.cell(row=T + e + 1, column=2).value = q[e]
            res.cell(row=T + e + 1, column=3).value = float(ApproxLin(tab_En, tab_Ga, E[e]) * q[e])
            res.cell(row=T + e + 1, column=7).value = ''
            b = 0.
            mud = np.zeros(len(dt))
            if E[e] > tab_E.max() or E[e] < tab_E.min():
                f = CubicSpline(tab_E, tab_mu[:, 0], extrapolate=True)
                if f(E[e]) > 0:
                    mus = f(E[e])
                else:
                    mus = ApproxPov(tab_E, tab_mu[:, 0], E[e])
            else:
                mus = ApproxLin(tab_E, tab_mu[:, 0], E[e])
            res.cell(row=T + e + 1, column=4).value = float(mus)
            print('μs =', mus)
            for i in range(len(dt)):
                if E[e] > tab_E.max() or E[e] < tab_E.min():
                    f = CubicSpline(tab_E, tab_mu[:, dt[i]], extrapolate=True)
                    if f(E[e]) > 0:
                        mud[i] = f(E[e])
                    else:
                        mud[i] = ApproxPov(tab_E, tab_mu[:, dt[i]], E[e])
                else:
                    mud[i] = ApproxLin(tab_E, tab_mu[:, dt[i]], E[e])
                res.cell(row=T + e + 1, column=7).value += ' μd' + str(i+1) + ' = ' + str(mud[i])
                print('μd[', i, '] =', mud[i])
                b += (mud[i] * d[i]) / 10
            res.cell(row=T + e + 1, column=9).value = b
            musH = float(mus * H * 100)
            res.cell(row=T + e + 1, column=8).value = musH
            s = np.zeros(len(tab_RH))
            Gpbk = np.zeros((len(tab_aH), len(tab_b1), len(tab_musH)))
            for i in range(len(tab_musH)):
                for j in range(len(tab_b1)):
                    for n in range(len(tab_aH)):
                        for m in range(len(tab_RH)):
                            s[m] = tab_Z[m, n, j, i]
                        Gpbk[n, j, i] = Approx(tab_RH, s, RH, 'log')
            s = np.zeros(len(tab_aH))
            Gpb = np.zeros((len(tab_b1), len(tab_musH)))
            for i in range(len(tab_musH)):
                for j in range(len(tab_b1)):
                    for n in range(len(tab_aH)):
                        s[n] = Gpbk[n, j, i]
                    Gpb[j, i] = Approx(tab_aH, s, aH, 'exp')
            s = np.zeros(len(tab_b1))
            Gp = np.zeros(len(tab_musH))
            for i in range(len(tab_musH)):
                for j in range(len(tab_b1)):
                    s[j] = Gpb[j, i]
                Gp[i] = Approx(tab_b1, s, b, 'exp')
            if musH > tab_musH[-2]:
                Z = float(Gp[-1] * np.exp(tab_musH[-2] * np.log(Gp[-2] / Gp[-1]) / musH))
            else:
                Z = ApproxLin(tab_musH, Gp, musH)
            res.cell(row=T + e + 1, column=13).value = Z
            V = H * np.pi * R ** 2
            D[e] = 2 * np.pi * A[e] * ApproxLin(tab_En, tab_Ga, E[e]) * q[e] * Z / (V * 100 * mus)
            res.cell(row=T + e + 1, column=15).value = D[e]
            print('μsH = ', musH)
            print('b = ', b)
            print('R/h = ', RH)
            print("a/h = ", aH)
            print("Z = ", Z)
            print('E = ', E[e])
            print('q = ', q[e])
            print('Г = ', float(ApproxLin(tab_En, tab_Ga, E[e]) * q[e]))
            print('V = ', V)
    print('Total dose rate, D = ', np.sum(D), 'μSv/h')
    res.cell(row=T, column=17).value = np.sum(D)
    T += rows + 1
results.save("Results"+datetime.now().strftime("%y%m%d%H%M%S")+".xlsx")